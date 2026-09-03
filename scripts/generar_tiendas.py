#!/usr/bin/env python3
"""
Genera data/tiendas.json a partir del reporte de ventas
"Central List / Account L4 / Route / Product Name" (12 semanas), tal cual se
exporta desde el sistema de ventas.

Uso:
    pip install openpyxl
    python3 scripts/generar_tiendas.py "Central_List....xlsx" [salida.json]

Qué hace:
  - Agrupa las filas (una por producto) en una fila por tienda.
  - Separa el campo "Sales Person" en número de ruta + nombre del dueño.
  - Suma las unidades por semana y calcula hace cuántas semanas fue la
    última con actividad — esto es lo que usa depuracion.html para avisar
    qué tiendas llevan tiempo sin visitarse.
  - Descarta la fila de "Total" que trae el reporte al final.

Vuelve a correr este script cada vez que haya un reporte nuevo; sobreescribe
data/tiendas.json. Los cambios que ya hicieron los IBPs (activa/inactiva/
frecuencia) NO se pierden — viven en el localStorage de cada quien y se
combinan con lo que sea que traiga este archivo por el "id" de la tienda.
"""

import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl. Instálalo con: pip install openpyxl")


def title(s):
    if not s:
        return s
    return " ".join(w.capitalize() for w in str(s).strip().split())


def generar(ruta_excel, hoja=None):
    wb = openpyxl.load_workbook(ruta_excel, read_only=True, data_only=True)
    ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]
    filas = list(ws.iter_rows(values_only=True))

    # El reporte trae varias líneas de metadata arriba; la fila de encabezado
    # real es la primera que empieza con celda vacía seguida de "Central Store".
    idx_header = next(
        i for i, r in enumerate(filas) if r and len(r) > 1 and r[1] == "Central Store"
    )
    header = filas[idx_header]
    weeklabels = list(header[11:23])
    datos = filas[idx_header + 1 :]

    tiendas = {}
    for r in datos:
        if not r or not r[1]:
            continue
        store_id = str(r[1]).strip()
        if not store_id or store_id.lower().startswith("total"):
            continue

        accountl4, salesperson = r[2], (r[3] or "").strip()
        store_name, addr, city, state = (r[5] or "").strip(), (r[6] or "").strip(), (r[7] or "").strip(), (r[8] or "").strip()
        zip_ = str(r[9]).strip() if r[9] is not None else ""
        semanas = [(w or 0) for w in r[11:23]]
        total = r[24] if len(r) > 24 else sum(semanas)

        m = re.match(r"^(\S+)\s+(.*)$", salesperson)
        ruta_num, propietario = (m.group(1), m.group(2).strip()) if m else (salesperson, "")

        if store_id not in tiendas:
            tiendas[store_id] = {
                "id": store_id,
                "ruta": ruta_num,
                "propietario": title(propietario),
                "nombre": title(store_name),
                "direccion": title(addr),
                "ciudad": title(city),
                "estado": state,
                "zip": zip_,
                "tipoCuenta": accountl4,
                "semanas": [0] * len(weeklabels),
                "totalUnidades": 0,
                "productos": 0,
            }
        st = tiendas[store_id]
        st["productos"] += 1
        for i, w in enumerate(semanas):
            st["semanas"][i] += w
        st["totalUnidades"] += total or 0

    for st in tiendas.values():
        con_actividad = [i for i, w in enumerate(st["semanas"]) if w != 0]
        st["semanasConActividad"] = len(con_actividad)
        if con_actividad:
            ultimo = max(con_actividad)
            st["ultimaSemanaConActividad"] = weeklabels[ultimo]
            st["semanasDesdeUltimaActividad"] = (len(st["semanas"]) - 1) - ultimo
        else:
            st["ultimaSemanaConActividad"] = None
            st["semanasDesdeUltimaActividad"] = len(st["semanas"])

    rutas = {}
    for st in tiendas.values():
        rutas.setdefault(st["ruta"], {"ruta": st["ruta"], "propietario": st["propietario"], "tiendas": []})
        rutas[st["ruta"]]["tiendas"].append(st)
    for r in rutas.values():
        r["tiendas"].sort(key=lambda s: (-s["semanasDesdeUltimaActividad"], s["nombre"]))

    from datetime import date

    return {
        "generadoEl": date.today().isoformat(),
        "rangoSemanas": {"desde": weeklabels[0], "hasta": weeklabels[-1], "etiquetas": weeklabels},
        "rutas": sorted(rutas.values(), key=lambda r: r["ruta"]),
    }


if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(__doc__)
    entrada = sys.argv[1]
    salida = sys.argv[2] if len(sys.argv) > 2 else str(Path(__file__).resolve().parent.parent / "data" / "tiendas.json")

    resultado = generar(entrada)
    Path(salida).write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8")

    total_tiendas = sum(len(r["tiendas"]) for r in resultado["rutas"])
    print(f"Listo: {salida}")
    print(f"{len(resultado['rutas'])} rutas, {total_tiendas} tiendas.")
